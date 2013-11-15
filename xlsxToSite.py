#!/bin/python

# key: fiEuEOjmjwnpLto0wT4KzA
# secret: F0RCUcgG59GO3uSP1yAmu0XKRKa7ZReJkpE3Kevr70

from openpyxl import load_workbook
import urllib2
from lxml import etree
import json
import os
import pystache

CWD = os.getcwd()
template = open(os.path.join(CWD, 'templates/index.html')).read()
outputFile = open(os.path.join(CWD, 'index.html'), 'w')
wb = load_workbook('Employee Free Book list for 11-13.xlsx')
# wb = load_workbook('test-11-13.xlsx')

worksheets = [
    wb.get_active_sheet(),
    wb.get_sheet_by_name('eBooks')
]

def generateJSONFromXLSX(ws):

    isbn_numbers = [isbn.value for isbn in ws.columns[0]][1:]
    titles = [title.value for title in ws.columns[1]][1:]
    format = [format.value for format in ws.columns[4]][1:]
    filter_format_options = []
    book_data = []

    for idx, isbn in enumerate(isbn_numbers):
        url = 'http://www.goodreads.com/book/isbn?format=XML&isbn=%s&key=fiEuEOjmjwnpLto0wT4KzA' % isbn
        try:
            response = urllib2.urlopen(url)
        except urllib2.HTTPError:
            print "SKIPPING %s" % isbn
            continue

        data = response.read()
        xml = etree.fromstring(data)
        book_details = {}
        book = xml.find('book')
        book_details['title'] = titles[idx]
        book_details['isbn'] = isbn
        book_details['format'] = format[idx]

        if format[idx] not in filter_format_options:
            filter_format_options.append(format[idx])

        book_details['cover_image'] = book.find('image_url').text
        book_details['rating'] = book.find('average_rating').text
        book_details['ratings_count'] = book.find('ratings_count').text
        book_details['description'] = book.find('description').text
        book_details['reviews_widget'] = book.find('reviews_widget').text
        book_details['url'] = book.find('url').text
        book_details['authors'] = [author.find('name').text for author in book.find('authors')]
        book_data.append(book_details)
        
    # for verbosity:
    print json.dumps(book_data, indent=4)
    return (book_data, filter_format_options)

# initialize

book_data = []
filter_format_options = []
 
for ws in worksheets:
    data = generateJSONFromXLSX(ws)
    book_data += data[0]
    filter_format_options += data[1]

context = {
    'books': book_data,
    'filter_format_options': filter_format_options
}

outputFile.write(pystache.render(template, context).encode('utf-8'))
outputFile.close()
